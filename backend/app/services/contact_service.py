"""Service for contact import from CSV/Excel."""

import csv
import io
import logging
from datetime import datetime
from typing import Dict, List, Tuple

from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from sqlalchemy.exc import IntegrityError

from app.models.contact import Contact

logger = logging.getLogger(__name__)


class ContactImportService:
    """Handles importing contacts from CSV and Excel files."""

    REQUIRED_FIELDS = {"email"}
    OPTIONAL_FIELDS = {"name", "phone", "city", "interests"}

    def __init__(self, db: AsyncSession):
        self.db = db

    async def import_csv(self, file_content: bytes) -> Dict[str, any]:
        """Import contacts from CSV file content."""
        stats = {"imported": 0, "skipped": 0, "errors": []}
        try:
            text = file_content.decode("utf-8")
            reader = csv.DictReader(io.StringIO(text))
            for row_num, row in enumerate(reader, start=2):
                await self._process_row(row, row_num, stats)
        except Exception as e:
            stats["errors"].append(f"Failed to parse CSV: {str(e)}")
        return stats

    async def import_excel(self, file_content: bytes) -> Dict[str, any]:
        """Import contacts from Excel file content."""
        stats = {"imported": 0, "skipped": 0, "errors": []}
        try:
            wb = load_workbook(io.BytesIO(file_content), read_only=True)
            ws = wb.active
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                row_dict = dict(zip(headers, row))
                await self._process_row(row_dict, row_num, stats)
        except Exception as e:
            stats["errors"].append(f"Failed to parse Excel: {str(e)}")
        return stats

    async def _process_row(self, row: Dict, row_num: int, stats: Dict) -> None:
        """Process a single row from import file."""
        email = row.get("email", "").strip() if row.get("email") else ""
        if not email:
            stats["errors"].append(f"Row {row_num}: missing email")
            stats["skipped"] += 1
            return

        # Check if contact exists
        result = await self.db.execute(select(Contact).where(Contact.email == email))
        existing = result.scalar_one_or_none()
        if existing:
            stats["skipped"] += 1
            return

        interests = row.get("interests")
        if isinstance(interests, str):
            interests = [i.strip() for i in interests.split(",") if i.strip()]

        contact = Contact(
            email=email,
            name=row.get("name", "").strip() if row.get("name") else None,
            phone=row.get("phone", "").strip() if row.get("phone") else None,
            city=row.get("city", "").strip() if row.get("city") else None,
            interests=interests if interests else None,
            consent_given=True,  # Imported with consent confirmation
            consent_date=datetime.utcnow(),
            source="import",
        )
        self.db.add(contact)
        try:
            await self.db.flush()
            stats["imported"] += 1
        except IntegrityError:
            await self.db.rollback()
            stats["skipped"] += 1

    async def export_contacts(self) -> List[Dict]:
        """Export all contacts as list of dicts for CSV/Excel export."""
        result = await self.db.execute(select(Contact))
        contacts = result.scalars().all()
        return [
            {
                "email": c.email,
                "name": c.name or "",
                "phone": c.phone or "",
                "city": c.city or "",
                "interests": ", ".join(c.interests) if c.interests else "",
                "is_subscribed": c.is_subscribed,
                "consent_given": c.consent_given,
            }
            for c in contacts
        ]
